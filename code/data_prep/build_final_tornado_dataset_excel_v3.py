#!/usr/bin/env python3

from pathlib import Path
import math
import warnings
import time

import geopandas as gpd
import numpy as np
import pandas as pd
from shapely.geometry import LineString, Point
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# =========================
# EDIT THESE PATHS
# =========================
CLEANED_CSV = r"us_tornado_cleaned_all_years.csv"
COUNTY_SNAPSHOT_CSV = r"acs_county_snapshot_2016_2020.csv"
COUNTY_SHP = r"tl_2020_us_county/tl_2020_us_county.shp"

OUTPUT_XLSX = r"final_tornado_dataset.xlsx"
OUTPUT_CSV = r"final_tornado_dataset.csv"
OUTPUT_BRIDGE_CSV = r"tornado_county_bridge.csv"
SAVE_BRIDGE_CSV = True

# =========================
# SETTINGS YOU CAN CHANGE
# =========================
COUNTY_CRS_PROJECTED = "EPSG:5070"   # contiguous US equal-area; good for area/overlap
COUNTY_CRS_WGS84 = "EPSG:4326"
WRITE_START_COUNTY_FEATURES = True
WRITE_END_COUNTY_IDS = True
USE_START_COUNTY_FALLBACK = False     # if True, fill path_* socioeconomic features from start county when no path buffer result exists

# If width is available, the tornado footprint is a line/point buffered by width/2.
# If width is missing, no path buffer is created. You can still use start-county features.

COUNTY_FEATURES = [
    "county_population",
    "county_population_density_km2",
    "county_mobile_home_share",
    "county_median_hh_income",
    "county_unemployment_rate",
    "county_disability_rate",
    "county_age_under18_share",
    "county_age_65plus_share",
]

PATH_WEIGHTED_OUTPUTS = {
    "county_population_density_km2": "path_population_density_wavg",
    "county_mobile_home_share": "path_mobile_home_share_wavg",
    "county_median_hh_income": "path_median_hh_income_wavg",
    "county_unemployment_rate": "path_unemployment_rate_wavg",
    "county_disability_rate": "path_disability_rate_wavg",
    "county_age_under18_share": "path_age_under18_share_wavg",
    "county_age_65plus_share": "path_age_65plus_share_wavg",
}

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN_GRAY = Side(style="thin", color="D9D9D9")


LOG_EVERY_ROWS = 5000


def log(msg: str):
    print(f"[{time.strftime('%H:%M:%S')}] {msg}", flush=True)



# =========================
# HELPERS
# =========================
def load_inputs():
    log("Loading cleaned tornado CSV...")
    tornado = pd.read_csv(CLEANED_CSV, low_memory=False)
    log(f"Loaded tornado rows: {len(tornado):,}")
    log("Loading county snapshot CSV...")
    county_snapshot = pd.read_csv(COUNTY_SNAPSHOT_CSV)
    log(f"Loaded county snapshot rows: {len(county_snapshot):,}")
    log("Loading county shapefile...")
    counties = gpd.read_file(COUNTY_SHP)
    log(f"Loaded county polygons: {len(counties):,}")
    return tornado, county_snapshot, counties


def prep_counties(county_snapshot: pd.DataFrame, counties: gpd.GeoDataFrame):
    log("Preparing county geometries and merging county features...")
    counties = counties.copy()

    if counties.crs is None:
        raise ValueError("County shapefile has no CRS. Re-download the TIGER/Line county shapefile.")

    if "GEOID" not in counties.columns:
        raise ValueError("County shapefile is missing GEOID.")
    if "ALAND" not in counties.columns:
        raise ValueError("County shapefile is missing ALAND.")

    county_snapshot = county_snapshot.copy()
    county_snapshot["GEOID"] = county_snapshot["GEOID"].astype(str).str.zfill(5)
    counties["GEOID"] = counties["GEOID"].astype(str).str.zfill(5)

    merged = counties.merge(county_snapshot, on="GEOID", how="left", suffixes=("", "_snap"))

    merged["county_population_density_per_m2"] = merged["county_population"] / merged["ALAND"]
    merged["county_population_density_km2"] = merged["county_population"] / (merged["ALAND"] / 1_000_000.0)

    required_after_merge = [
        "county_population",
        "county_population_density_per_m2",
        "county_population_density_km2",
        "county_mobile_home_share",
        "county_median_hh_income",
        "county_unemployment_rate",
        "county_disability_rate",
        "county_age_under18_share",
        "county_age_65plus_share",
    ]
    missing_after_merge = [c for c in required_after_merge if c not in merged.columns]
    if missing_after_merge:
        raise ValueError(f"County merge is missing expected columns: {missing_after_merge}")

    log("Projecting county geometries to WGS84 and projected CRS...")
    counties_wgs = merged.to_crs(COUNTY_CRS_WGS84)
    counties_proj = merged.to_crs(COUNTY_CRS_PROJECTED)
    log("County preparation done.")

    return counties_wgs, counties_proj


def ensure_tornado_id(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    if "clean_row_id" in df.columns:
        df["tornado_id"] = df["clean_row_id"].astype(int)
    elif "raw_row_id" in df.columns:
        df["tornado_id"] = df["raw_row_id"].astype(int)
    else:
        df["tornado_id"] = np.arange(len(df), dtype=int)
    return df


def valid_latlon(lat, lon):
    return pd.notna(lat) and pd.notna(lon) and (-90 <= lat <= 90) and (-180 <= lon <= 180)


def build_start_end_point_tables(df: pd.DataFrame):
    log("Building start/end point tables...")
    start_rows = []
    end_rows = []

    total = len(df)
    for i, (_, row) in enumerate(df.iterrows(), start=1):
        tid = int(row["tornado_id"])

        if valid_latlon(row.get("slat"), row.get("slon")):
            start_rows.append({
                "tornado_id": tid,
                "geometry": Point(float(row["slon"]), float(row["slat"])),
            })

        if valid_latlon(row.get("elat"), row.get("elon")):
            end_rows.append({
                "tornado_id": tid,
                "geometry": Point(float(row["elon"]), float(row["elat"])),
            })

        if i % LOG_EVERY_ROWS == 0 or i == total:
            log(f"  Start/end points processed: {i:,}/{total:,}")

    start_gdf = gpd.GeoDataFrame(start_rows, geometry="geometry", crs=COUNTY_CRS_WGS84)
    end_gdf = gpd.GeoDataFrame(end_rows, geometry="geometry", crs=COUNTY_CRS_WGS84)
    log(f"Built start points: {len(start_gdf):,}; end points: {len(end_gdf):,}")
    return start_gdf, end_gdf


def spatial_join_points(points_gdf: gpd.GeoDataFrame, counties_wgs: gpd.GeoDataFrame, prefix: str):
    log(f"Spatially joining {prefix} points to counties...")
    if points_gdf.empty:
        return pd.DataFrame(columns=["tornado_id", f"{prefix}_county_geoid", f"{prefix}_county_name"])

    county_cols = ["GEOID", "NAME"] + [c for c in COUNTY_FEATURES if c in counties_wgs.columns]
    join_cols = counties_wgs[county_cols + ["geometry"]].copy()

    joined = gpd.sjoin(points_gdf, join_cols, how="left", predicate="within")
    keep = ["tornado_id", "GEOID", "NAME"] + [c for c in COUNTY_FEATURES if c in joined.columns]
    joined = joined[keep].copy()
    joined = joined.rename(columns={
        "GEOID": f"{prefix}_county_geoid",
        "NAME": f"{prefix}_county_name",
    })

    rename_map = {c: f"{prefix}_{c}" for c in COUNTY_FEATURES if c in joined.columns}
    joined = joined.rename(columns=rename_map)
    joined = joined.drop_duplicates(subset=["tornado_id"])
    log(f"  {prefix} county matches: {len(joined):,}")
    return pd.DataFrame(joined)


def build_tornado_buffers(df: pd.DataFrame):
    log("Building tornado path buffers...")
    source_rows = []

    total = len(df)
    for i, (_, row) in enumerate(df.iterrows(), start=1):
        tid = int(row["tornado_id"])
        slat, slon = row.get("slat"), row.get("slon")
        elat, elon = row.get("elat"), row.get("elon")
        wid = row.get("wid")

        if not valid_latlon(slat, slon):
            continue
        if pd.isna(wid) or float(wid) <= 0:
            continue

        width_m = float(wid) * 0.9144
        radius_m = width_m / 2.0

        has_end = valid_latlon(elat, elon)
        same_point = has_end and math.isclose(float(slat), float(elat), rel_tol=0.0, abs_tol=1e-10) and math.isclose(float(slon), float(elon), rel_tol=0.0, abs_tol=1e-10)

        if has_end and not same_point:
            geom = LineString([(float(slon), float(slat)), (float(elon), float(elat))])
            method = "corridor_buffer"
        else:
            geom = Point(float(slon), float(slat))
            method = "start_point_buffer"

        source_rows.append({
            "tornado_id": tid,
            "buffer_radius_m": radius_m,
            "county_overlay_method": method,
            "geometry": geom,
        })

        if i % LOG_EVERY_ROWS == 0 or i == total:
            log(f"  Buffer candidates processed: {i:,}/{total:,}")

    if not source_rows:
        return gpd.GeoDataFrame(columns=["tornado_id", "buffer_radius_m", "county_overlay_method", "geometry"], geometry="geometry", crs=COUNTY_CRS_WGS84)

    log(f"Projecting and buffering {len(source_rows):,} tornado geometries...")
    base = gpd.GeoDataFrame(source_rows, geometry="geometry", crs=COUNTY_CRS_WGS84).to_crs(COUNTY_CRS_PROJECTED)
    base["geometry"] = [geom.buffer(rad) for geom, rad in zip(base.geometry, base["buffer_radius_m"])]
    log("Finished building tornado buffers.")
    return base


def build_path_overlap_features(buffers_gdf: gpd.GeoDataFrame, counties_proj: gpd.GeoDataFrame):
    log("Computing tornado-county path overlaps...")
    if buffers_gdf.empty:
        empty_agg = pd.DataFrame(columns=[
            "tornado_id",
            "county_overlay_method",
            "path_counties_n",
            "path_overlap_area_m2",
            "path_overlap_area_km2",
            "path_est_exposed_pop",
        ] + list(PATH_WEIGHTED_OUTPUTS.values()))
        empty_bridge = pd.DataFrame()
        return empty_agg, empty_bridge

    counties_lookup = counties_proj[["GEOID", "NAME", "ALAND", "county_population_density_per_m2"] + [c for c in COUNTY_FEATURES if c in counties_proj.columns] + ["geometry"]].copy()
    counties_lookup = counties_lookup.reset_index(drop=True)
    counties_lookup["county_idx"] = counties_lookup.index

    counties_join = counties_lookup[["county_idx", "geometry"]].copy()
    counties_join = gpd.GeoDataFrame(counties_join, geometry="geometry", crs=COUNTY_CRS_PROJECTED)

    log(f"Running spatial join for {len(buffers_gdf):,} buffers against {len(counties_proj):,} counties...")
    candidates = gpd.sjoin(
        buffers_gdf[["tornado_id", "buffer_radius_m", "county_overlay_method", "geometry"]],
        counties_join,
        how="inner",
        predicate="intersects",
    ).reset_index(drop=True)

    log(f"Candidate buffer-county intersections: {len(candidates):,}")

    if candidates.empty:
        empty_agg = pd.DataFrame(columns=[
            "tornado_id",
            "county_overlay_method",
            "path_counties_n",
            "path_overlap_area_m2",
            "path_overlap_area_km2",
            "path_est_exposed_pop",
        ] + list(PATH_WEIGHTED_OUTPUTS.values()))
        empty_bridge = pd.DataFrame()
        return empty_agg, empty_bridge

    candidates = candidates.merge(
        counties_lookup.rename(columns={"geometry": "county_geometry", "NAME": "county_name"}),
        on="county_idx",
        how="left",
    )

    overlap_areas = []
    total_candidates = len(candidates)
    for i, (geom_left, geom_right) in enumerate(zip(candidates["geometry"], candidates["county_geometry"]), start=1):
        try:
            overlap_areas.append(geom_left.intersection(geom_right).area)
        except Exception:
            overlap_areas.append(np.nan)

        if i % LOG_EVERY_ROWS == 0 or i == total_candidates:
            log(f"  Overlap areas computed: {i:,}/{total_candidates:,}")
    candidates["overlap_area_m2"] = overlap_areas
    candidates = candidates[candidates["overlap_area_m2"].fillna(0) > 0].copy()

    if candidates.empty:
        empty_agg = pd.DataFrame(columns=[
            "tornado_id",
            "county_overlay_method",
            "path_counties_n",
            "path_overlap_area_m2",
            "path_overlap_area_km2",
            "path_est_exposed_pop",
        ] + list(PATH_WEIGHTED_OUTPUTS.values()))
        empty_bridge = pd.DataFrame()
        return empty_agg, empty_bridge

    log(f"Positive overlap rows: {len(candidates):,}")
    candidates["estimated_exposed_pop"] = candidates["county_population_density_per_m2"] * candidates["overlap_area_m2"]

    # bridge output (optional, good for auditing)
    bridge_cols = [
        "tornado_id",
        "county_overlay_method",
        "GEOID",
        "county_name",
        "overlap_area_m2",
        "estimated_exposed_pop",
    ] + [c for c in COUNTY_FEATURES if c in candidates.columns]
    bridge = candidates[bridge_cols].copy().rename(columns={"GEOID": "county_geoid"})

    # aggregated path metrics
    group = candidates.groupby("tornado_id", dropna=False)
    agg = group.agg(
        county_overlay_method=("county_overlay_method", "first"),
        path_counties_n=("GEOID", "nunique"),
        path_overlap_area_m2=("overlap_area_m2", "sum"),
        path_est_exposed_pop=("estimated_exposed_pop", "sum"),
    ).reset_index()
    agg["path_overlap_area_km2"] = agg["path_overlap_area_m2"] / 1_000_000.0

    log("Aggregating weighted path features...")
    for source_col, output_col in PATH_WEIGHTED_OUTPUTS.items():
        tmp = candidates[["tornado_id", "estimated_exposed_pop", source_col]].copy()
        tmp = tmp[tmp[source_col].notna()].copy()
        if tmp.empty:
            agg[output_col] = np.nan
            continue
        tmp["weighted_val"] = tmp["estimated_exposed_pop"] * tmp[source_col]
        numer = tmp.groupby("tornado_id")["weighted_val"].sum().rename("numer")
        denom = tmp.groupby("tornado_id")["estimated_exposed_pop"].sum().rename("denom")
        merged = pd.concat([numer, denom], axis=1).reset_index()
        merged[output_col] = np.where(merged["denom"] > 0, merged["numer"] / merged["denom"], np.nan)
        agg = agg.merge(merged[["tornado_id", output_col]], on="tornado_id", how="left")

    log(f"Built aggregated path features for {len(agg):,} tornado rows.")
    return agg, bridge


def append_fallbacks(final_df: pd.DataFrame) -> pd.DataFrame:
    final_df = final_df.copy()

    # Keep explicit flags so you can choose later which rows/features to trust.
    final_df["path_features_present"] = final_df["path_counties_n"].notna().astype("int8")
    final_df["start_county_present"] = final_df["start_county_geoid"].notna().astype("int8")

    if USE_START_COUNTY_FALLBACK:
        fill_map = {
            "path_population_density_wavg": "start_county_county_population_density_km2",
            "path_mobile_home_share_wavg": "start_county_county_mobile_home_share",
            "path_median_hh_income_wavg": "start_county_county_median_hh_income",
            "path_unemployment_rate_wavg": "start_county_county_unemployment_rate",
            "path_disability_rate_wavg": "start_county_county_disability_rate",
            "path_age_under18_share_wavg": "start_county_county_age_under18_share",
            "path_age_65plus_share_wavg": "start_county_county_age_65plus_share",
        }
        no_path = final_df["path_features_present"] == 0
        for path_col, start_col in fill_map.items():
            if path_col in final_df.columns and start_col in final_df.columns:
                final_df.loc[no_path & final_df[start_col].notna(), path_col] = final_df.loc[no_path & final_df[start_col].notna(), start_col]
        final_df.loc[no_path & final_df["start_county_present"].eq(1), "county_overlay_method"] = final_df.loc[no_path & final_df["start_county_present"].eq(1), "county_overlay_method"].fillna("start_county_fallback")

    final_df["path_overlay_or_start_fallback_present"] = (
        final_df["path_features_present"].eq(1)
        | (
            USE_START_COUNTY_FALLBACK
            and final_df["start_county_present"].eq(1)
        )
    ).astype("int8")

    return final_df


def build_dictionary_sheet_columns(df: pd.DataFrame):
    rows = []
    new_cols = [
        ("start_county_geoid", "County GEOID containing the tornado start point."),
        ("start_county_name", "County name containing the tornado start point."),
        ("end_county_geoid", "County GEOID containing the tornado end point, when end coordinates exist."),
        ("end_county_name", "County name containing the tornado end point, when end coordinates exist."),
        ("county_overlay_method", "How path-based county features were built: corridor_buffer, start_point_buffer, or fallback/blank."),
        ("path_counties_n", "Number of counties intersected by the buffered tornado footprint."),
        ("path_overlap_area_m2", "Total area of tornado-buffer overlap with counties, in square meters."),
        ("path_overlap_area_km2", "Total area of tornado-buffer overlap with counties, in square kilometers."),
        ("path_est_exposed_pop", "Estimated exposed population from county density × overlap area."),
        ("path_population_density_wavg", "Population-weighted average county population density across intersected counties."),
        ("path_mobile_home_share_wavg", "Population-weighted average mobile-home share across intersected counties."),
        ("path_median_hh_income_wavg", "Population-weighted average median household income across intersected counties."),
        ("path_unemployment_rate_wavg", "Population-weighted average unemployment rate across intersected counties."),
        ("path_disability_rate_wavg", "Population-weighted average disability rate across intersected counties."),
        ("path_age_under18_share_wavg", "Population-weighted average under-18 share across intersected counties."),
        ("path_age_65plus_share_wavg", "Population-weighted average 65+ share across intersected counties."),
        ("path_features_present", "1 if path-based county features were successfully computed, else 0."),
        ("start_county_present", "1 if the tornado start point could be assigned to a county, else 0."),
        ("path_overlay_or_start_fallback_present", "1 if path features exist, or if start-county fallback filled them (when enabled)."),
    ]
    for col, desc in new_cols:
        if col in df.columns:
            rows.append([col, desc])

    if WRITE_START_COUNTY_FEATURES:
        for col in df.columns:
            if col.startswith("start_county_county_"):
                rows.append([col, "County snapshot feature taken directly from the county containing the tornado start point."])

    return pd.DataFrame(rows, columns=["column_name", "description"])


def write_excel(final_df: pd.DataFrame, dictionary_df: pd.DataFrame, bridge_df: pd.DataFrame | None):
    log("Writing Excel workbook...")
    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        final_df.to_excel(writer, index=False, sheet_name="final_dataset")
        dictionary_df.to_excel(writer, index=False, sheet_name="data_dictionary")
        if SAVE_BRIDGE_CSV and bridge_df is not None and not bridge_df.empty:
            bridge_preview = bridge_df.head(25000).copy()
            bridge_preview.to_excel(writer, index=False, sheet_name="bridge_preview")

    log("Applying Excel formatting...")
    wb = load_workbook(OUTPUT_XLSX)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.sheet_view.showGridLines = False

        max_col = ws.max_column
        max_row = ws.max_row

        # Header styling
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = Border(bottom=THIN_GRAY)

        # Add table when size is reasonable
        if max_row >= 2 and max_col >= 1 and max_row <= 1_048_576:
            ref = f"A1:{get_column_letter(max_col)}{max_row}"
            table = Table(displayName=f"tbl_{ws.title[:20].replace(' ', '_')}", ref=ref)
            style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            table.tableStyleInfo = style
            ws.add_table(table)

        # Widths / number formats
        for col_idx, col_cells in enumerate(ws.columns, start=1):
            header = ws.cell(1, col_idx).value
            width = max(len(str(header)) if header is not None else 0, 12)
            for cell in col_cells[1:min(max_row, 300)]:
                if cell.value is not None:
                    width = min(max(width, len(str(cell.value)) + 2), 28)

            if isinstance(header, str) and ("date" in header.lower()):
                for cell in col_cells[1:]:
                    cell.number_format = "yyyy-mm-dd"
            elif isinstance(header, str) and ("share" in header.lower() or "rate" in header.lower()):
                for cell in col_cells[1:]:
                    cell.number_format = "0.0%"
            elif isinstance(header, str) and ("income" in header.lower()):
                for cell in col_cells[1:]:
                    cell.number_format = "$#,##0"
            elif isinstance(header, str) and ("density" in header.lower() or "area" in header.lower() or "pop" in header.lower()):
                for cell in col_cells[1:]:
                    cell.number_format = "0.00"

            ws.column_dimensions[get_column_letter(col_idx)].width = width

    wb.save(OUTPUT_XLSX)
    log("Excel workbook saved.")


def main():
    warnings.filterwarnings("ignore", category=UserWarning)
    t0 = time.perf_counter()
    log("Starting final tornado dataset build...")

    tornado, county_snapshot, counties = load_inputs()
    tornado = ensure_tornado_id(tornado)

    counties_wgs, counties_proj = prep_counties(county_snapshot, counties)

    # Start / end county assignment
    start_pts, end_pts = build_start_end_point_tables(tornado)
    start_join = spatial_join_points(start_pts, counties_wgs, "start") if WRITE_START_COUNTY_FEATURES else pd.DataFrame(columns=["tornado_id"])
    end_join = spatial_join_points(end_pts, counties_wgs, "end") if WRITE_END_COUNTY_IDS else pd.DataFrame(columns=["tornado_id"])

    # Path-based county features
    buffers = build_tornado_buffers(tornado)
    path_agg, bridge = build_path_overlap_features(buffers, counties_proj)

    log("Merging county features back into tornado table...")
    final_df = tornado.merge(start_join, on="tornado_id", how="left")
    final_df = final_df.merge(end_join, on="tornado_id", how="left")
    final_df = final_df.merge(path_agg, on="tornado_id", how="left")
    final_df = append_fallbacks(final_df)

    # Reorder so appended fields go to the end
    original_cols = [c for c in tornado.columns]
    appended_cols = [c for c in final_df.columns if c not in original_cols]
    final_df = final_df[original_cols + appended_cols]

    # Save flat outputs
    log("Writing final CSV...")
    final_df.to_csv(OUTPUT_CSV, index=False)
    if SAVE_BRIDGE_CSV and not bridge.empty:
        log("Writing bridge CSV...")
        bridge.to_csv(OUTPUT_BRIDGE_CSV, index=False)

    # Excel workbook
    dictionary_df = build_dictionary_sheet_columns(final_df)
    write_excel(final_df, dictionary_df, bridge if SAVE_BRIDGE_CSV else None)

    elapsed = time.perf_counter() - t0
    print("Saved final CSV:", Path(OUTPUT_CSV).resolve())
    print("Saved final Excel:", Path(OUTPUT_XLSX).resolve())
    if SAVE_BRIDGE_CSV and not bridge.empty:
        print("Saved bridge CSV:", Path(OUTPUT_BRIDGE_CSV).resolve())
    print("Rows:", len(final_df))
    print("Columns:", len(final_df.columns))
    print("Path features present:", int(final_df["path_features_present"].sum()))
    print("Start county present:", int(final_df["start_county_present"].sum()))
    print(f"Elapsed seconds: {elapsed:,.1f}")


if __name__ == "__main__":
    main()
